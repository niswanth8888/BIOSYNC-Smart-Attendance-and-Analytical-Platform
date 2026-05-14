import os
from datetime import datetime
from collections import defaultdict

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference, PieChart

# ==================================================
# CONFIG
# ==================================================

BACKUP_FOLDER = r"C:\AttendanceAutomation\BackupExcel"
OUTPUT_FOLDER = r"C:\AttendanceAutomation\RangeReports"
LOGO_PATH = r"C:\AttendanceAutomation\logo.jpg"

DEPARTMENT = None        # "AIML"
CLASS_NAME = "I ECE E"       # "III AIML"      
EMP_CODE = "25106281"          #"23148033"

START_DATE = "2026-03-01"
END_DATE   = "2026-03-24"

# ==================================================
# STYLES
# ==================================================

thin = Side(style="thin")
border = Border(left=thin,right=thin,top=thin,bottom=thin)

def style_table(ws, sr, er, ec):
    for r in range(sr, er+1):
        for c in range(1, ec+1):
            cell = ws.cell(r,c)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            if r == sr:
                cell.font = Font(bold=True)

def auto_width(ws):
    for col in ws.columns:
        l = 0
        letter = col[0].column_letter
        for cell in col:
            if cell.value:
                l = max(l, len(str(cell.value)))
        ws.column_dimensions[letter].width = l + 4

def add_footer(ws, label):
    fr = ws.max_row + 3
    try:
        img = Image(LOGO_PATH)
        img.width = 120
        img.height = 60
        ws.add_image(img, f"A{fr}")
    except:
        pass

    tr = fr + 4
    ws.merge_cells(start_row=tr,start_column=1,end_row=tr,end_column=4)

    t = ws.cell(tr,1)
    t.value = "BIOSYNC Attendance Engine"
    t.font = Font(bold=True,size=12)
    t.alignment = Alignment(horizontal="center")

    ws.cell(tr+1,1).value = f"Class : {label}"
    ws.cell(tr+2,1).value = f"Range : {START_DATE} to {END_DATE}"
    ws.cell(tr+3,1).value = f"Generated : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

# ==================================================
# ⭐ PROFESSIONAL CHART FIX
# ==================================================

def set_professional_chart(bar, total_items):
    bar.style = 10   # closest to Excel Style-9
    bar.width = max(25, total_items * 0.45)
    bar.height = 18

# ==================================================
# DATE FILTER
# ==================================================

sd = datetime.strptime(START_DATE,"%Y-%m-%d").date()
ed = datetime.strptime(END_DATE,"%Y-%m-%d").date()

files = []

for m in os.listdir(BACKUP_FOLDER):
    mp = os.path.join(BACKUP_FOLDER,m)
    if not os.path.isdir(mp): continue

    for f in os.listdir(mp):
        if not f.endswith(".xlsx"): continue
        try:
            d = datetime.strptime(f.replace(".xlsx",""),"%d-%m-%Y").date()
        except:
            continue
        if d.weekday()==6: continue
        if sd <= d <= ed:
            files.append((d, os.path.join(mp,f)))

files.sort()

wb_out = Workbook()

# ==================================================
# STUDENT MODE
# ==================================================

if EMP_CODE:

    ws = wb_out.active
    ws.title = "Report"

    ws["A1"] = "EMP CODE"
    ws["B1"] = EMP_CODE
    ws["A2"] = "Class"
    ws["B2"] = CLASS_NAME

    row = 4
    ws.append(["Date","Name","Punch","Status"])

    start = row
    present = 0
    total = 0
    name = ""

    for d,p in files:
        wb = load_workbook(p,data_only=True)
        if CLASS_NAME not in wb.sheetnames:
            wb.close(); continue

        sh = wb[CLASS_NAME]
        total += 1
        found = False

        for r in range(2, sh.max_row+1):
            if str(sh.cell(r,2).value) == str(EMP_CODE):
                name = sh.cell(r,4).value
                punch = sh.cell(r,5).value
                status = sh.cell(r,10).value or "Absent"

                ws.append([d.strftime("%d-%m-%Y"),name,punch,status])

                if status=="Present": present+=1
                found=True
                break

        if not found:
            ws.append([d.strftime("%d-%m-%Y"),name,"-","Absent"])

        wb.close()

    end = ws.max_row
    style_table(ws,start,end,4)

    ws.append([])
    ws.append(["SUMMARY"])

    absent = total - present
    pct = round((present/total)*100,2) if total else 0

    srow = ws.max_row+1
    ws.append(["Working Days",total])
    ws.append(["Present Days",present])
    ws.append(["Absent Days",absent])
    ws.append(["Attendance %",pct])

    style_table(ws,srow,ws.max_row,2)

    add_footer(ws, CLASS_NAME)
    auto_width(ws)

    ch = wb_out.create_sheet("Analytics")

    ch.append(["Status","Count"])
    ch.append(["Present",present])
    ch.append(["Absent",absent])

    pie = PieChart()
    data = Reference(ch,min_col=2,min_row=1,max_row=3)
    labels = Reference(ch,min_col=1,min_row=2,max_row=3)
    pie.add_data(data,titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Attendance Split"

    ch.add_chart(pie,"D4")

# ==================================================
# CLASS MODE
# ==================================================

elif CLASS_NAME and not DEPARTMENT:

    ws = wb_out.active
    ws.title = CLASS_NAME

    row = 1
    stats = defaultdict(lambda: {"p":0,"t":0})

    for d,p in files:
        wb = load_workbook(p,data_only=True)
        if CLASS_NAME not in wb.sheetnames:
            wb.close(); continue

        sh = wb[CLASS_NAME]

        ws.cell(row,1).value = d.strftime("%d-%m-%Y")
        ws.cell(row,1).font = Font(bold=True,size=12)
        row+=1

        start = row
        ws.append(["Emp","Name","Punch","Status"])

        for r in range(2, sh.max_row+1):
            emp = sh.cell(r,2).value
            if not emp: continue

            name = sh.cell(r,4).value
            punch = sh.cell(r,5).value
            status = sh.cell(r,10).value or "Absent"

            ws.append([emp,name,punch,status])

            stats[name]["t"]+=1
            if status=="Present":
                stats[name]["p"]+=1

        style_table(ws,start,ws.max_row,4)
        row = ws.max_row + 2
        wb.close()

    ws.cell(row,1).value = "CLASS SUMMARY"
    ws.cell(row,1).font = Font(bold=True,size=12)
    row+=2

    start = row
    ws.append(["Student","%","Present","Absent"])

    class_pct = []

    for n,d in stats.items():
        p = d["p"]
        t = d["t"]
        a = t-p
        pct = round((p/t)*100,2) if t else 0
        ws.append([n,pct,p,a])
        class_pct.append(pct)

    style_table(ws,start,ws.max_row,4)

    avg = round(sum(class_pct)/len(class_pct),2) if class_pct else 0

    add_footer(ws, CLASS_NAME)
    auto_width(ws)

    ch = wb_out.create_sheet("Analytics")

    ch.append(["Student","%"])
    for n,d in stats.items():
        pct = round((d["p"]/d["t"])*100,2) if d["t"] else 0
        ch.append([n,pct])

    bar = BarChart()
    data = Reference(ch,min_col=2,min_row=1,max_row=ch.max_row)
    cats = Reference(ch,min_col=1,min_row=2,max_row=ch.max_row)
    bar.add_data(data,titles_from_data=True)
    bar.set_categories(cats)
    bar.title = "Student Attendance %"

    set_professional_chart(bar, ch.max_row)

    ch.add_chart(bar,"D4")

    ch.append([])
    ch.append(["Class Avg",avg])

# ==================================================
# DEPARTMENT MODE
# ==================================================

elif DEPARTMENT:

    class_avg = {}

    for d,p in files:

        wb = load_workbook(p,data_only=True)

        for s in wb.sheetnames:

            if DEPARTMENT not in s:
                continue

            if s not in wb_out.sheetnames:
                ws = wb_out.create_sheet(s)
                row = 1
            else:
                ws = wb_out[s]
                row = ws.max_row + 2

            sh = wb[s]

            ws.cell(row,1).value = d.strftime("%d-%m-%Y")
            ws.cell(row,1).font = Font(bold=True,size=12)

            row += 1

            start = row
            ws.append(["Emp","Name","Punch","Status"])

            for r in range(2, sh.max_row+1):

                emp = sh.cell(r,2).value
                if not emp:
                    continue

                ws.append([
                    emp,
                    sh.cell(r,4).value,
                    sh.cell(r,5).value,
                    sh.cell(r,10).value
                ])

            style_table(ws,start,ws.max_row,4)

        wb.close()

    for ws in wb_out.worksheets:

        if ws.title == "Sheet":
            continue

        stats = defaultdict(lambda:{"p":0,"t":0})

        for r in range(1, ws.max_row+1):

            name = ws.cell(r,2).value
            status = ws.cell(r,4).value

            if status in ("Present","Absent","Not Present"):

                stats[name]["t"] += 1
                if status == "Present":
                    stats[name]["p"] += 1

        ws.append([])
        ws.append(["Student","%","Present","Absent"])

        start = ws.max_row

        pct_list = []

        for n,d in stats.items():

            p = d["p"]
            t = d["t"]
            a = t - p

            pct = round((p/t)*100,2) if t else 0

            ws.append([n,pct,p,a])
            pct_list.append(pct)

        style_table(ws,start,ws.max_row,4)

        avg = round(sum(pct_list)/len(pct_list),2) if pct_list else 0
        class_avg[ws.title] = avg

        add_footer(ws, ws.title)
        auto_width(ws)

    dash = wb_out.create_sheet("Department Dashboard")

    dash.append(["Class","Attendance %"])

    for c,p in class_avg.items():
        dash.append([c,p])

    bar = BarChart()

    data = Reference(dash,min_col=2,min_row=1,max_row=dash.max_row)
    cats = Reference(dash,min_col=1,min_row=2,max_row=dash.max_row)

    bar.add_data(data,titles_from_data=True)
    bar.set_categories(cats)
    bar.title = "Department Comparison"

    set_professional_chart(bar, dash.max_row)

    dash.add_chart(bar,"E5")

# ==================================================
# SAVE
# ==================================================

if "Sheet" in wb_out.sheetnames:
    del wb_out["Sheet"]

os.makedirs(OUTPUT_FOLDER,exist_ok=True)

name = EMP_CODE or CLASS_NAME or DEPARTMENT
fn = f"{name}_{START_DATE}_to_{END_DATE}.xlsx"

wb_out.save(os.path.join(OUTPUT_FOLDER,fn))

print("✅ BIOSYNC RANGE REPORT GENERATED")