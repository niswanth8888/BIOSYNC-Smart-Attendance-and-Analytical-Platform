from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from openpyxl import load_workbook
import json
import os
import subprocess
from datetime import datetime
from collections import defaultdict

app = FastAPI()

# ================= CORS =================

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ================= BASE PATH =================

BASE_DIR = r"C:\biosync_web_app"
SCRIPT_DIR = os.path.join(BASE_DIR, "scripts")
CONFIG_DIR = os.path.join(BASE_DIR, "config")

REPORT_CONFIG = os.path.join(CONFIG_DIR, "report_config.json")
SDK_CONFIG = os.path.join(CONFIG_DIR, "sdk_config.json")
SDK_LAST_STATUS = os.path.join(CONFIG_DIR, "sdk_last_status.json")
DEPARTMENT_USERS_FILE = os.path.join(BASE_DIR, "department_users.json")

ATTENDANCE_MASTER_FILE = r"C:\AttendanceAutomation\attendance_master.xlsx"
BACKUP_FOLDER = r"C:\AttendanceAutomation\BackupExcel"

# ================= MODELS =================

class LoginData(BaseModel):
    username: str
    password: str

class DepartmentRangeReportRequest(BaseModel):
    department: str
    start_date: str
    end_date: str

# ================= DEPARTMENT SHEET MAP =================

DEPARTMENT_SHEETS = {
    "CSE": [
        "I CSE A", "I CSE B", "I CSE C",
        "II CSE A", "II CSE B", "II CSE C",
        "III CSE A", "III CSE B", "III CSE C",
        "IV CSE A", "IV CSE B", "IV CSE C"
    ],
    "CSBS": [
        "I CSBS", "II CSBS", "III CSBS", "IV CSBS"
    ],
    "AIML": [
        "I AIML A", "I AIML B", "II AIML A", "II AIML B", "III AIML"
    ],
    "IT": [
        "I IT A", "I IT B", "I IT C",
        "II IT A", "II IT B", "II IT C",
        "III IT A", "III IT B", "III IT C",
        "IV IT A", "IV IT B", "IV IT C"
    ],
    "MECH": [
        "I MECH", "II MECH", "III MECH", "IV MECH"
    ],
    "CIVIL": [
        "I CIVIL", "II CIVIL", "III CIVIL", "IV CIVIL"
    ],
    "CHEMICAL": [
        "I CHEMICAL", "II CHEMICAL", "III CHEMICAL", "IV CHEMICAL"
    ],
    "EEE": [
        "I EEE", "II EEE", "III EEE", "IV EEE"
    ],
    "CCE": [
        "I CCE", "II CCE", "III CCE", "IV CCE"
    ],
    "BIOTECH": [
        "I BIOTECH", "II BIOTECH", "III BIOTECH", "IV BIOTECH"
    ],
    "BME": [
        "I BME", "II BME", "III BME", "IV BME"
    ],
    "AIDS": [
        "I AIDS A", "I AIDS B", "I AIDS C", "I AIDS D",
        "II AIDS A", "II AIDS B", "II AIDS C", "II AIDS D",
        "III AIDS A", "III AIDS B",
        "IV AIDS A", "IV AIDS B"
    ],
    "ECE": [
        "I ECE A", "I ECE B", "I ECE C", "I ECE D", "I ECE E", "I ECE F",
        "II ECE A", "II ECE B", "II ECE C", "II ECE D",
        "III ECE A", "III ECE B", "III ECE C", "III ECE D",
        "IV ECE A", "IV ECE B", "IV ECE C"
    ]
}

# ================= HELPERS =================

def safe_percentage(present: int, total: int) -> float:
    if total <= 0:
        return 0.0
    return round((present / total) * 100, 2)

def is_present_value(value) -> bool:
    if value is None:
        return False
    return str(value).strip().lower() == "present"

def get_no_cache_headers():
    return {
        "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
        "Pragma": "no-cache",
        "Expires": "0"
    }

def read_json_file(file_path, default_value):
    if not os.path.exists(file_path):
        return default_value

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return default_value

def build_sdk_status_payload():
    default_status = {
        "successful_devices": [],
        "offline_devices": [],
        "no_log_devices": [],
        "error_devices": [],
        "last_run": None
    }

    if not os.path.exists(SDK_LAST_STATUS):
        return default_status

    try:
        with open(SDK_LAST_STATUS, "r", encoding="utf-8") as f:
            data = json.load(f)

        return {
            "successful_devices": data.get("successful_devices", []),
            "offline_devices": data.get("offline_devices", []),
            "no_log_devices": data.get("no_log_devices", []),
            "error_devices": data.get("error_devices", []),
            "last_run": data.get("last_run")
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read SDK last status: {str(e)}")

def get_existing_department_sheet_map(workbook):
    existing_sheetnames = set(workbook.sheetnames)
    filtered_map = {}

    for dept, sheets in DEPARTMENT_SHEETS.items():
        existing = [sheet for sheet in sheets if sheet in existing_sheetnames]
        filtered_map[dept] = existing

    return filtered_map

def count_sheet_attendance(ws):
    present = 0
    absent = 0
    total = 0

    for row in range(2, ws.max_row + 1):
        student_value = ws.cell(row, 2).value
        status_value = ws.cell(row, 10).value

        if student_value is None or str(student_value).strip() == "":
            continue

        total += 1

        if is_present_value(status_value):
            present += 1
        else:
            absent += 1

    return {
        "present": present,
        "absent": absent,
        "total": total
    }

def sum_department_attendance(workbook, department_name: str):
    dept_key = department_name.strip().upper()
    department_map = get_existing_department_sheet_map(workbook)

    if dept_key not in department_map:
        raise HTTPException(status_code=404, detail=f"Department '{department_name}' not found")

    present = 0
    absent = 0
    total = 0

    for sheet_name in department_map[dept_key]:
        ws = workbook[sheet_name]
        sheet_data = count_sheet_attendance(ws)
        present += sheet_data["present"]
        absent += sheet_data["absent"]
        total += sheet_data["total"]

    return {
        "name": dept_key,
        "present": present,
        "absent": absent,
        "total": total,
        "attendance_percentage": safe_percentage(present, total)
    }

def sum_college_attendance(workbook):
    department_map = get_existing_department_sheet_map(workbook)

    present = 0
    absent = 0
    total = 0
    counted_sheets = set()

    for _, sheets in department_map.items():
        for sheet_name in sheets:
            if sheet_name in counted_sheets:
                continue

            counted_sheets.add(sheet_name)
            ws = workbook[sheet_name]
            sheet_data = count_sheet_attendance(ws)
            present += sheet_data["present"]
            absent += sheet_data["absent"]
            total += sheet_data["total"]

    return {
        "present": present,
        "absent": absent,
        "total": total
    }

def build_department_comparison(workbook, requested_department: str):
    department_map = get_existing_department_sheet_map(workbook)
    comparison = []

    for dept_name in department_map.keys():
        dept_data = sum_department_attendance(workbook, dept_name)
        comparison.append(dept_data)

    requested_key = requested_department.strip().upper()

    comparison.sort(
        key=lambda item: (
            0 if item["name"] == requested_key else 1,
            item["name"]
        )
    )

    return comparison

def parse_input_date(date_str: str):
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Date format must be YYYY-MM-DD")

def find_backup_files_in_range(start_date, end_date):
    if not os.path.exists(BACKUP_FOLDER):
        raise HTTPException(status_code=500, detail="BackupExcel folder missing")

    files = []

    for month_folder in os.listdir(BACKUP_FOLDER):
        month_path = os.path.join(BACKUP_FOLDER, month_folder)
        if not os.path.isdir(month_path):
            continue

        for filename in os.listdir(month_path):
            if not filename.lower().endswith(".xlsx"):
                continue

            try:
                file_date = datetime.strptime(filename.replace(".xlsx", ""), "%d-%m-%Y").date()
            except Exception:
                continue

            if file_date.weekday() == 6:
                continue

            if start_date <= file_date <= end_date:
                files.append((file_date, os.path.join(month_path, filename)))

    files.sort(key=lambda x: x[0])
    return files

def get_existing_sheets_for_department_in_workbook(workbook, department_name: str):
    dept_key = department_name.strip().upper()
    if dept_key not in DEPARTMENT_SHEETS:
        raise HTTPException(status_code=404, detail=f"Department '{department_name}' not found")

    return [sheet for sheet in DEPARTMENT_SHEETS[dept_key] if sheet in workbook.sheetnames]

def count_present_total_for_sheet(ws):
    present = 0
    total = 0

    for row in range(2, ws.max_row + 1):
        student_value = ws.cell(row, 2).value
        status_value = ws.cell(row, 10).value

        if student_value is None or str(student_value).strip() == "":
            continue

        total += 1
        if is_present_value(status_value):
            present += 1

    return present, total

def build_department_range_report(department: str, start_date_str: str, end_date_str: str):
    dept_key = department.strip().upper()
    start_date = parse_input_date(start_date_str)
    end_date = parse_input_date(end_date_str)

    if start_date > end_date:
        raise HTTPException(status_code=400, detail="start_date cannot be greater than end_date")

    if dept_key not in DEPARTMENT_SHEETS:
        raise HTTPException(status_code=404, detail=f"Department '{department}' not found")

    files = find_backup_files_in_range(start_date, end_date)

    class_stats = {
        class_name: {
            "present_sum": 0,
            "total_sum": 0,
            "days_found": 0,
            "student_counts": []
        }
        for class_name in DEPARTMENT_SHEETS[dept_key]
    }

    daily_trend = []

    for file_date, file_path in files:
        try:
            wb = load_workbook(file_path, data_only=True)
        except Exception:
            continue

        existing_class_sheets = get_existing_sheets_for_department_in_workbook(wb, dept_key)

        day_present = 0
        day_total = 0

        for class_name in existing_class_sheets:
            ws = wb[class_name]
            present, total = count_present_total_for_sheet(ws)

            class_stats[class_name]["present_sum"] += present
            class_stats[class_name]["total_sum"] += total
            class_stats[class_name]["days_found"] += 1
            class_stats[class_name]["student_counts"].append(total)

            day_present += present
            day_total += total

        if day_total > 0:
            daily_trend.append({
                "date": file_date.strftime("%Y-%m-%d"),
                "attendance_percentage": safe_percentage(day_present, day_total)
            })

        wb.close()

    filtered_class_stats = {
        class_name: stats
        for class_name, stats in class_stats.items()
        if stats["total_sum"] > 0
    }

    total_present_all = sum(stats["present_sum"] for stats in filtered_class_stats.values())
    total_possible_all = sum(stats["total_sum"] for stats in filtered_class_stats.values())
    overall_average = safe_percentage(total_present_all, total_possible_all)

    class_comparison = []
    class_breakdown = []

    for class_name, stats in filtered_class_stats.items():
        attendance_percentage = safe_percentage(stats["present_sum"], stats["total_sum"])

        if stats["student_counts"]:
            total_students = max(stats["student_counts"])
        else:
            total_students = 0

        present_average = round(
            stats["present_sum"] / stats["days_found"], 2
        ) if stats["days_found"] > 0 else 0.0

        class_comparison.append({
            "class_name": class_name,
            "attendance_percentage": attendance_percentage
        })

        class_breakdown.append({
            "class_name": class_name,
            "total_students": total_students,
            "present_average": present_average,
            "attendance_percentage": attendance_percentage
        })

    class_comparison.sort(key=lambda x: x["class_name"])
    class_breakdown.sort(key=lambda x: x["class_name"])
    daily_trend.sort(key=lambda x: x["date"])

    return {
        "department": dept_key,
        "start_date": start_date_str,
        "end_date": end_date_str,
        "overall_average": overall_average,
        "class_comparison": class_comparison,
        "daily_trend": daily_trend,
        "class_breakdown": class_breakdown
    }

# ================= ADMIN LOGIN =================

@app.post("/admin/login")
def admin_login(data: LoginData):

    file_path = os.path.join(BASE_DIR, "admin_users.json")

    if not os.path.exists(file_path):
        raise HTTPException(status_code=500, detail="admin_users.json missing")

    with open(file_path, "r", encoding="utf-8") as f:
        admins = json.load(f)

    for admin in admins:
        if admin["username"] == data.username and admin["password"] == data.password:
            return {"status": "success"}

    raise HTTPException(status_code=401, detail="Invalid login")

# ================= DEPARTMENT LOGIN =================

@app.post("/department/login")
def department_login(data: LoginData):

    if not os.path.exists(DEPARTMENT_USERS_FILE):
        raise HTTPException(status_code=500, detail="department_users.json missing")

    try:
        with open(DEPARTMENT_USERS_FILE, "r", encoding="utf-8") as f:
            users = json.load(f)

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read department_users.json: {str(e)}")

    for user in users:
        if user.get("username") == data.username and user.get("password") == data.password:
            return {
                "status": "success",
                "role": "department",
                "department": user.get("department"),
                "display_name": user.get("display_name", user.get("department"))
            }

    raise HTTPException(status_code=401, detail="Invalid department login")

# ================= DEPARTMENT DASHBOARD SUMMARY =================

@app.get("/department/dashboard-summary/{department}")
def get_department_dashboard_summary(department: str):

    if not os.path.exists(ATTENDANCE_MASTER_FILE):
        raise HTTPException(status_code=500, detail="attendance_master.xlsx missing")

    try:
        wb = load_workbook(ATTENDANCE_MASTER_FILE, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to open attendance master file: {str(e)}")

    try:
        college_data = sum_college_attendance(wb)
        department_data = sum_department_attendance(wb, department)
        comparison_data = build_department_comparison(wb, department)

        return {
            "date": datetime.now().strftime("%Y-%m-%d"),
            "college": college_data,
            "department": department_data,
            "comparison": comparison_data
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to build department dashboard summary: {str(e)}")
    finally:
        wb.close()

# ================= DEPARTMENT RANGE REPORT =================

@app.post("/department/range-report")
def department_range_report(data: DepartmentRangeReportRequest):
    try:
        return build_department_range_report(
            department=data.department,
            start_date_str=data.start_date,
            end_date_str=data.end_date
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to build department range report: {str(e)}")

# ================= SDK CONFIG =================

@app.get("/sdk/config")
def get_sdk_config():
    headers = get_no_cache_headers()
    config_data = read_json_file(SDK_CONFIG, {})
    return JSONResponse(content=config_data, headers=headers)

@app.put("/sdk/config")
def update_sdk_config(data: dict):

    os.makedirs(CONFIG_DIR, exist_ok=True)

    with open(SDK_CONFIG, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4)

    return {"status": "updated"}

# ================= SDK LAST STATUS =================

@app.get("/sdk/last-status")
def get_sdk_last_status():
    headers = get_no_cache_headers()
    status_payload = build_sdk_status_payload()
    return JSONResponse(content=status_payload, headers=headers)

# ================= SDK BOOTSTRAP =================

@app.get("/sdk/bootstrap")
def get_sdk_bootstrap():
    headers = get_no_cache_headers()

    config_data = read_json_file(SDK_CONFIG, {})
    status_payload = build_sdk_status_payload()

    return JSONResponse(
        content={
            "config": config_data,
            "status": status_payload
        },
        headers=headers
    )

# ================= RUN SDK =================

@app.post("/run/sdk_pull")
def run_sdk():

    script = os.path.join(SCRIPT_DIR, "sdk_pull_attendance.py")

    if not os.path.exists(script):
        raise HTTPException(status_code=500, detail="SDK script missing")

    subprocess.Popen(
        f'start cmd /k python "{script}"',
        shell=True
    )

    return {"status": "sdk started"}

# ================= RUN COMBINE =================

@app.post("/run/combine")
def run_combine():

    script = os.path.join(SCRIPT_DIR, "combine_all_sheets.py")

    if not os.path.exists(script):
        raise HTTPException(status_code=500, detail="Combine script missing")

    subprocess.Popen(
        f'start cmd /k python "{script}"',
        shell=True
    )

    return {"status": "combine started"}

# ================= REPORT CONFIG =================

@app.get("/report/config")
def get_report_config():

    if not os.path.exists(REPORT_CONFIG):
        return {
            "department": None,
            "class_name": None,
            "emp_code": None,
            "start_date": None,
            "end_date": None
        }

    with open(REPORT_CONFIG, "r", encoding="utf-8") as f:
        return json.load(f)

@app.put("/report/config")
def update_report_config(data: dict):

    os.makedirs(CONFIG_DIR, exist_ok=True)

    config = {
        "department": data.get("department") or None,
        "class_name": data.get("class_name") or None,
        "emp_code": data.get("emp_code") or None,
        "start_date": data.get("start_date"),
        "end_date": data.get("end_date")
    }

    with open(REPORT_CONFIG, "w", encoding="utf-8") as f:
        json.dump(config, f, indent=4)

    return {
        "status": "config_saved",
        "config": config
    }

# ================= RUN RANGE REPORT =================

@app.post("/run/report")
def run_report():

    script = os.path.join(SCRIPT_DIR, "biosync_range_report.py")

    if not os.path.exists(script):
        raise HTTPException(status_code=500, detail="Report engine missing")

    subprocess.Popen(
        f'start cmd /k python "{script}"',
        shell=True
    )

    return {"status": "report started"}