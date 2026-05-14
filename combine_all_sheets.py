from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side

SOURCE_FILE = r"C:\AttendanceAutomation\attendance_master.xlsx"
OUTPUT_FILE = r"C:\AttendanceAutomation\Input\attendance_combined.xlsx"

# ===============================
# LOAD WORKBOOKS
# ===============================
wb_src = load_workbook(SOURCE_FILE)
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "Employee Punch Monitor"

# ===============================
# STYLES (EXACT)
# ===============================
bold = Font(bold=True)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")

border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

row_ptr = 1

# ===============================
# TITLE
# ===============================
ws_out.cell(row=row_ptr, column=1, value="Employee Punch Monitor").font = Font(bold=True, size=14)
row_ptr += 2

ws_out.cell(row=row_ptr, column=1, value="Company:")
ws_out.cell(row=row_ptr, column=2, value="VSBEC")
ws_out.cell(row=row_ptr, column=1).font = bold
row_ptr += 2

# ===============================
# PROCESS EACH DEPARTMENT (IN ORDER)
# ===============================
for sheet_name in wb_src.sheetnames:
    ws = wb_src[sheet_name]

    # -------------------------------
    # DEPARTMENT ROW
    # -------------------------------
    ws_out.cell(row=row_ptr, column=1, value="Department").font = bold
    ws_out.cell(row=row_ptr, column=4, value=sheet_name).font = bold

    for col in range(1, 11):
        ws_out.cell(row=row_ptr, column=col).border = border

    row_ptr += 1

    # -------------------------------
    # HEADER ROW (EXACT POSITIONS)
    # -------------------------------
    headers = {
        1: "SNo.",
        2: "Emp Code",
        4: "Name",
        5: "Last Punch",
        7: "Direction",
        8: "Punch Records",
        10: "Status"
    }

    for col, text in headers.items():
        c = ws_out.cell(row=row_ptr, column=col, value=text)
        c.font = bold
        c.alignment = center
        c.border = border

    for col in [3, 6, 9]:
        ws_out.cell(row=row_ptr, column=col).border = border

    row_ptr += 1
    sno = 1

    # -------------------------------
    # AUTO-DETECT FIRST DATA ROW
    # -------------------------------
    start_row = None
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 2).value).strip().isdigit():
            start_row = r
            break

    if not start_row:
        row_ptr += 2
        continue

    # -------------------------------
    # COPY DATA (NO SKIPPING)
    # -------------------------------
    for r in range(start_row, ws.max_row + 1):
        emp_code = ws.cell(r, 2).value
        name = ws.cell(r, 4).value

        if not emp_code:
            continue

        values = {
            1: sno,
            2: emp_code,
            4: name,
            5: ws.cell(r, 5).value,
            7: ws.cell(r, 7).value,
            8: ws.cell(r, 8).value,
            10: ws.cell(r, 10).value
        }

        for col, val in values.items():
            c = ws_out.cell(row=row_ptr, column=col, value=val)
            c.border = border
            c.alignment = center if col in [1, 2, 5] else left

        for col in [3, 6, 9]:
            ws_out.cell(row=row_ptr, column=col).border = border

        sno += 1
        row_ptr += 1

    row_ptr += 2  # space between departments

# ===============================
# SAVE
# ===============================
wb_out.save(OUTPUT_FILE)

print("✅ PERFECT COMBINED FILE CREATED")
print(f"📂 File Location: {OUTPUT_FILE}")
