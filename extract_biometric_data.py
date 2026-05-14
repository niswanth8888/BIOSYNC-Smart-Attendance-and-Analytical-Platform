import http.server
import socketserver
import threading
import queue
import time
import re
from datetime import datetime
from openpyxl import load_workbook

# ==================================================
# CONFIGURATION
# ==================================================

EXCEL_FILE = r"C:\AttendanceAutomation\attendance_master.xlsx"
DEVICE_LOG_FILE = r"C:\AttendanceAutomation\device_status.log"

PORT_DEVICE_MAP = {
    9937: "I AIML A",
    8945: "I AIML B",
    65500: "III AIML",
    2601: "I CSE A",
    6523: "I CSE B",
    6701: "I CSE C",
    1132: "II CSE A",
    3726: "II CSE B",
    4027: "II CSE C",
    1257: "III CSE A",
    4056: "III CSE B",
    5501: "III CSE C",
    4670: "IV CSE A",
    9825: "IV CSE B",
    6275: "IV CSE C",
    4253: "I CSBS",
    7727: "II CSBS",
    8939: "III CSBS",
    7827: "IV CSBS",
    5011: "I IT A",
    5012: "I IT B",
    5013: "I IT C",
    5014: "II IT A",
    5015: "II IT B",
    5016: "II IT C",
    5017: "III IT A",
    5018: "III IT B",
    5019: "III IT C",
    6001: "IV IT A",
    6002: "IV IT B",
    5022: "IV IT C",
}

LIVE_TIME_WINDOW = 120        # seconds
DEVICE_TIMEOUT = 30           # seconds

# ==================================================
# GLOBALS
# ==================================================

excel_lock = threading.Lock()
punch_queue = queue.Queue()

last_seen_punch = {}          # sheet -> emp -> HH:MM
device_last_seen = {}         # port -> timestamp
device_status = {}            # port -> ONLINE/OFFLINE

# ==================================================
# HELPERS
# ==================================================

def valid_punch(parts):
    return (
        len(parts) >= 3
        and parts[0].isdigit()
        and re.match(r"\d{4}-\d{2}-\d{2}", parts[1])
        and re.match(r"\d{2}:\d{2}:\d{2}", parts[2])
    )

def is_live_punch(date_str, time_str):
    try:
        punch_dt = datetime.strptime(
            f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S"
        )
        return abs((datetime.now() - punch_dt).total_seconds()) <= LIVE_TIME_WINDOW
    except:
        return False

# ==================================================
# DEVICE STATUS LOGGER (FIXED SIZE)
# ==================================================

def update_device_log():
    while True:
        lines = []
        now = time.time()

        for port, name in PORT_DEVICE_MAP.items():
            last = device_last_seen.get(port)
            status = "ONLINE" if last and (now - last <= DEVICE_TIMEOUT) else "OFFLINE"
            device_status[port] = status

            last_seen_str = (
                datetime.fromtimestamp(last).strftime("%Y-%m-%d %H:%M:%S")
                if last else "N/A"
            )

            lines.append(f"{port} | {name} | {status} | {last_seen_str}")

        with open(DEVICE_LOG_FILE, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))

        time.sleep(10)

# ==================================================
# EXCEL WRITER (FAST + SAFE)
# ==================================================

def excel_worker():
    while True:
        sheet, emp, punch = punch_queue.get()

        with excel_lock:
            wb = load_workbook(EXCEL_FILE)
            if sheet not in wb.sheetnames:
                wb.close()
                punch_queue.task_done()
                continue

            ws = wb[sheet]

            # ✅ FIXED: start from ROW 2 (NO SKIPPED STUDENTS)
            for row in range(2, ws.max_row + 1):
                emp_code = str(ws.cell(row, 2).value).strip()
                if emp_code == emp:
                    ws.cell(row, 5).value = punch
                    ws.cell(row, 8).value = (ws.cell(row, 8).value or "") + punch + ","
                    ws.cell(row, 10).value = "Present"
                    break

            wb.save(EXCEL_FILE)
            wb.close()

        print(f"💾 SAVED → {sheet} | {emp} | {punch}")
        punch_queue.task_done()

# ==================================================
# HTTP HANDLER (SAFE ADMS)
# ==================================================

class BioHandler(http.server.BaseHTTPRequestHandler):

    def _ack(self):
        self.send_response(200)
        self.end_headers()
        self.wfile.write(b"OK")

    def do_POST(self):
        port = self.server.server_address[1]
        sheet = PORT_DEVICE_MAP.get(port)

        length = int(self.headers.get("Content-Length", 0))
        payload = self.rfile.read(length).decode(errors="ignore")

        self._ack()   # ACK IMMEDIATELY — DO NOT BLOCK DEVICE

        if not sheet:
            return

        device_last_seen[port] = time.time()

        if sheet not in last_seen_punch:
            last_seen_punch[sheet] = {}

        for line in payload.splitlines():
            parts = line.split()
            if not valid_punch(parts):
                continue

            if not is_live_punch(parts[1], parts[2]):
                continue  # ignore buffered data

            emp = parts[0]
            punch = parts[2][:5]

            if last_seen_punch[sheet].get(emp) == punch:
                continue

            last_seen_punch[sheet][emp] = punch
            punch_queue.put((sheet, emp, punch))

        print(f"📥 LIVE DATA OK → {sheet}")

    def do_GET(self): self._ack()
    def do_HEAD(self): self._ack()
    def do_OPTIONS(self): self._ack()

    def log_message(self, *args):
        return

# ==================================================
# SERVER
# ==================================================

class ReusableServer(socketserver.ThreadingTCPServer):
    allow_reuse_address = True

def start_server(port, sheet):
    with ReusableServer(("", port), BioHandler) as server:
        print(f"🚀 Listening on PORT {port} → {sheet}")
        server.serve_forever()

# ==================================================
# MAIN
# ==================================================

if __name__ == "__main__":

    print("\n🚀 BIOSYNC LIVE-ONLY ATTENDANCE ENGINE (SAFE MODE)")
    print(f"📘 Excel File: {EXCEL_FILE}")
    print("=" * 70)

    threading.Thread(target=excel_worker, daemon=True).start()
    threading.Thread(target=update_device_log, daemon=True).start()

    for port, sheet in PORT_DEVICE_MAP.items():
        threading.Thread(
            target=start_server,
            args=(port, sheet),
            daemon=True
        ).start()

    threading.Event().wait()