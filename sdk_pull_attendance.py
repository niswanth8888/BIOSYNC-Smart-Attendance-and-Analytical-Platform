from zk import ZK
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
import threading
import socket
import os
import zipfile
import tempfile
import shutil

# ==================================================
# CONFIG
# ==================================================

EXCEL_FILE = r"C:\AttendanceAutomation\attendance_master.xlsx"
PAST_EXCEL_FILE = r"C:\AttendanceAutomation\attendance_master_past.xlsx"
DEVICE_LOG = r"C:\AttendanceAutomation\sdk_device_status.log"
LOGO_PATH = r"C:\AttendanceAutomation\logo.jpg"

TARGET_DEVICES = ["III BME","I CIVIL&CHEMICAL","III EEE","III ECE D","II EEE","IV EEE"]
AUTO_CLEAR_LOGS = False
FETCH_DATE = None
# ==================================================
# DEVICE LIST
# ==================================================
# 🔴 PASTE YOUR FULL DEVICE LIST HERE

DEVICES = [
 {"name": "I CSE A", "ip": "172.16.31.108",  "port": 4370, "sheet": "I CSE A"},
    {"name": "I CSE B", "ip": "172.16.31.109", "port": 4370, "sheet": "I CSE B"},
    {"name": "I CSE C", "ip": "172.16.31.110", "port": 4370, "sheet": "I CSE C"},
    {"name": "II CSE A", "ip": "172.16.31.111",  "port": 4370, "sheet": "II CSE A"},
    {"name": "II CSE B", "ip": "172.16.31.112", "port": 4370, "sheet": "II CSE B"},
    {"name": "II CSE C", "ip": "172.16.31.113", "port": 4370, "sheet": "II CSE C"},
    {"name": "III CSE A", "ip": "172.16.31.114",  "port": 4370, "sheet": "III CSE A"},
    {"name": "III CSE B", "ip": "172.16.31.115", "port": 4370, "sheet": "III CSE B"},
    {"name": "III CSE C", "ip": "172.16.31.116", "port": 4370, "sheet": "III CSE C"},
    {"name": "IV CSE A", "ip": "172.16.31.117",  "port": 4370, "sheet": "IV CSE A"},
    {"name": "IV CSE B", "ip": "172.16.31.118", "port": 4370, "sheet": "IV CSE B"},
    {"name": "IV CSE C", "ip": "172.16.31.119", "port": 4370, "sheet": "IV CSE C"},
    {"name": "I CSBS", "ip": "172.16.31.104",  "port": 4370, "sheet": "I CSBS"},
    {"name": "II CSBS", "ip": "172.16.31.105", "port": 4370, "sheet": "II CSBS"},
    {"name": "III CSBS", "ip": "172.16.31.106", "port": 4370, "sheet": "III CSBS"},
    {"name": "IV CSBS", "ip": "172.16.31.107",  "port": 4370, "sheet": "IV CSBS"},
    {"name": "I AIML A", "ip": "172.16.31.100", "port": 4370, "sheet": "I AIML A"},
    {"name": "I AIML B", "ip": "172.16.31.101", "port": 4370, "sheet": "I AIML B"},
    {"name": "II AIML A&B", "ip": "172.16.31.102", "port": 4370, "sheet": ["II AIML A", "II AIML B"]},
    {"name": "III AIML", "ip": "172.16.31.103", "port": 4370, "sheet": "III AIML"},
    {"name": "I IT A", "ip": "172.16.31.120",  "port": 4370, "sheet": "I IT A"},
    {"name": "I IT B", "ip": "172.16.31.121", "port": 4370, "sheet": "I IT B"},
    {"name": "I IT C", "ip": "172.16.31.122", "port": 4370, "sheet": "I IT C"},
    {"name": "II IT A", "ip": "172.16.31.123",  "port": 4370, "sheet": "II IT A"},
    {"name": "II IT B", "ip": "172.16.31.124", "port": 4370, "sheet": "II IT B"},
    {"name": "II IT C", "ip": "172.16.31.125", "port": 4370, "sheet": "II IT C"},
    {"name": "III IT A", "ip": "172.16.31.126",  "port": 4370, "sheet": "III IT A"},
    {"name": "III IT B", "ip": "172.16.31.127", "port": 4370, "sheet": "III IT B"},
    {"name": "III IT C", "ip": "172.16.31.128", "port": 4370, "sheet": "III IT C"},
    {"name": "IV IT A", "ip": "172.16.31.129",  "port": 4370, "sheet": "IV IT A"},
    {"name": "IV IT B&C", "ip": "172.16.31.131", "port": 4370, "sheet":["IV IT B","IV IT C"]},
    {"name": "I MECH", "ip": "172.16.31.132",  "port": 4370, "sheet": "I MECH"},
    {"name": "II MECH", "ip": "172.16.31.133", "port": 4370, "sheet": "II MECH"},
    {"name": "III MECH", "ip": "172.16.31.134", "port": 4370, "sheet": "III MECH"},
    {"name": "IV MECH", "ip": "172.16.31.135",  "port": 4370, "sheet": "IV MECH"},
    {"name": "I CIVIL&CHEMICAL", "ip": "172.16.31.136",  "port": 4370, "sheet": ["I CIVIL","I CHEMICAL"]},
    {"name": "II CIVIL&CHEMICAL", "ip": "172.16.31.137", "port": 4370, "sheet": ["II CIVIL","II CHEMICAL"]},
    {"name": "III CIVIL&CHEMICAL", "ip": "172.16.31.158", "port": 4370, "sheet": ["III CIVIL","III CHEMICAL"]},
    {"name": "IV CIVIL&CHEMICAL", "ip": "172.16.31.159",  "port": 4370, "sheet": ["IV CIVIL","IV CHEMICAL"]},
    {"name": "I EEE", "ip": "172.16.31.140",  "port": 4370, "sheet": "I EEE"},
    {"name": "II EEE", "ip": "172.16.31.141", "port": 4370, "sheet": "II EEE"},
    {"name": "III EEE", "ip": "172.16.31.142", "port": 4370, "sheet": "III EEE"},
    {"name": "IV EEE", "ip": "172.16.31.143",  "port": 4370, "sheet": "IV EEE"},
    {"name": "I CCE", "ip": "172.16.31.144",  "port": 4370, "sheet": "I CCE"},
    {"name": "II CCE", "ip": "172.16.31.145", "port": 4370, "sheet": "II CCE"},
    {"name": "III&IV CCE", "ip": "172.16.31.146", "port": 4370, "sheet": ["III CCE","IV CCE"]},
    {"name": "II BIOTECH", "ip": "172.16.31.149", "port": 4370, "sheet": "II BIOTECH"},
    {"name": "III BIOTECH", "ip": "172.16.31.150", "port": 4370, "sheet": "III BIOTECH"},
    {"name": "IV BIOTECH", "ip": "172.16.31.151",  "port": 4370, "sheet": "IV BIOTECH"},
    {"name": "I BME&BIOTECH", "ip": "172.16.31.148",  "port": 4370, "sheet": ["I BME","I BIOTECH"]},
    {"name": "II BME", "ip": "172.16.31.153", "port": 4370, "sheet": "II BME"},
    {"name": "III BME", "ip": "172.16.31.154", "port": 4370, "sheet": "III BME"},
    {"name": "IV BME", "ip": "172.16.31.155",  "port": 4370, "sheet": "IV BME"},
    {"name": "I AIDS A", "ip": "172.16.31.160",  "port": 4370, "sheet": "I AIDS A"},
    {"name": "I AIDS B", "ip": "172.16.31.161", "port": 4370, "sheet": "I AIDS B"},
    {"name": "I AIDS C", "ip": "172.16.31.162", "port": 4370, "sheet": "I AIDS C"},
    {"name": "I AIDS D", "ip": "172.16.31.163",  "port": 4370, "sheet": "I AIDS D"},
    {"name": "II AIDS A&B", "ip": "172.16.31.164", "port": 4370, "sheet": ["II AIDS A","II AIDS B"]},
    {"name": "II AIDS C&D", "ip": "172.16.31.166",  "port": 4370, "sheet": ["II AIDS C","II AIDS D"]},
    {"name": "III AIDS A", "ip": "172.16.31.168", "port": 4370, "sheet": "III AIDS A"},
    {"name": "III AIDS B", "ip": "172.16.31.169",  "port": 4370, "sheet": "III AIDS B"},
    {"name": "IV AIDS A&B", "ip": "172.16.31.170", "port": 4370, "sheet": ["IV AIDS A","IV AIDS B"]},
    {"name": "I ECE A", "ip": "172.16.31.172",  "port": 4370, "sheet": "I ECE A"},
    {"name": "I ECE B", "ip": "172.16.31.173", "port": 4370, "sheet": "I ECE B"},
    {"name": "I ECE C", "ip": "172.16.31.174", "port": 4370, "sheet": "I ECE C"},
    {"name": "I ECE D", "ip": "172.16.31.175",  "port": 4370, "sheet": "I ECE D"},
    {"name": "I ECE E", "ip": "172.16.31.176", "port": 4370, "sheet": "I ECE E"},
    {"name": "I ECE F", "ip": "172.16.31.177", "port": 4370, "sheet": "I ECE F"},
    {"name": "II ECE A", "ip": "172.16.31.178",  "port": 4370, "sheet": "II ECE A"},
    {"name": "II ECE B", "ip": "172.16.31.179", "port": 4370, "sheet": "II ECE B"},
    {"name": "II ECE C&D", "ip": "172.16.31.180", "port": 4370, "sheet": ["II ECE C","II ECE D"]},
    {"name": "III ECE A&B", "ip": "172.16.31.182", "port": 4370, "sheet": ["III ECE A","III ECE B"]},
    {"name": "III ECE C", "ip": "172.16.31.184", "port": 4370, "sheet": "III ECE C"},
    {"name": "III ECE D", "ip": "172.16.31.185", "port": 4370, "sheet": "III ECE D"},
    {"name": "IV ECE A", "ip": "172.16.31.186", "port": 4370, "sheet": "IV ECE A"},
    {"name": "IV ECE B", "ip": "172.16.31.187", "port": 4370, "sheet": "IV ECE B"},
    {"name": "IV ECE C", "ip": "172.16.31.188", "port": 4370, "sheet": "IV ECE C"},
]

# ==================================================
# DATE
# ==================================================

TODAY = date.today()

# ==================================================
# MODE SELECTION
# ==================================================

if FETCH_DATE:
    try:
        FETCH_DATE_OBJ = datetime.strptime(FETCH_DATE, "%Y-%m-%d").date()
        ACTIVE_EXCEL_FILE = PAST_EXCEL_FILE
        PAST_MODE = True
        print("📂 Running in PAST MODE")
    except ValueError:
        print("❌ Invalid FETCH_DATE format")
        exit()
else:
    FETCH_DATE_OBJ = TODAY
    ACTIVE_EXCEL_FILE = EXCEL_FILE
    PAST_MODE = False
    print("📂 Running in NORMAL MODE")

excel_lock = threading.Lock()
tracker_lock = threading.Lock()

# ==================================================
# EXECUTION TRACKERS
# ==================================================

successful_devices = []
offline_devices = []
no_log_devices = []
error_devices = []
clear_failed_devices = []

# ==================================================
# HELPERS
# ==================================================

def normalize_emp(val):
    if val is None:
        return ""
    if isinstance(val, float):
        val = int(val)
    return str(val).strip().lstrip("0")

def log_device(name, ip, status):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(DEVICE_LOG, "a", encoding="utf-8") as f:
        f.write(f"[{ts}] {name} ({ip}) → {status}\n")

def is_ip_reachable(ip):
    try:
        socket.create_connection((ip, 4370), timeout=3)
        return True
    except:
        return False

# ==================================================
# EXCEL COMPRESSION
# ==================================================

def compress_excel(file_path):

    temp_dir = tempfile.mkdtemp()

    with zipfile.ZipFile(file_path, 'r') as zin:
        zin.extractall(temp_dir)

    new_file = file_path.replace(".xlsx", "_compressed.xlsx")

    with zipfile.ZipFile(new_file,
                         'w',
                         compression=zipfile.ZIP_DEFLATED,
                         compresslevel=9) as zout:

        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                full = os.path.join(root, file)
                arc = os.path.relpath(full, temp_dir)
                zout.write(full, arc)

    shutil.move(new_file, file_path)
    shutil.rmtree(temp_dir)

# ==================================================
# UPDATE EXCEL
# ==================================================

def update_excel(sheet, records):

    with excel_lock:

        if not os.path.exists(ACTIVE_EXCEL_FILE):
            print("⚠ Excel file missing")
            return 0

        wb = load_workbook(ACTIVE_EXCEL_FILE)
        total_updated = 0

        if not isinstance(sheet, list):
            sheet = [sheet]

        for sh in sheet:

            if sh not in wb.sheetnames:
                continue

            ws = wb[sh]

            # Remove old footer
            for row in range(ws.max_row, 0, -1):
                if ws.cell(row, 1).value == "BIOSYNC Attendance Engine":
                    ws.delete_rows(row - 4, 15)
                    break

            # Remove old logos
            try:
                if hasattr(ws, "_images"):
                    ws._images.clear()
            except:
                pass

            # Clear past mode
            if PAST_MODE:
                for r in range(2, ws.max_row + 1):
                    if ws.cell(r, 2).value is None:
                        break
                    ws.cell(r, 5).value = None
                    ws.cell(r, 8).value = None
                    ws.cell(r, 10).value = None

            # Map students
            emp_map = {}
            for r in range(2, ws.max_row + 1):
                emp = ws.cell(r, 2).value
                if emp:
                    emp_map[normalize_emp(emp)] = r

            # Update attendance
            for emp, punch in records:

                emp_n = normalize_emp(emp)

                if emp_n in emp_map:

                    row = emp_map[emp_n]

                    ws.cell(row, 5).value = punch

                    existing = ws.cell(row, 8).value

                    if existing:
                        punch_list = [p for p in existing.split(",") if p.strip()]
                    else:
                        punch_list = []

                    if punch not in punch_list:
                        punch_list.append(punch)

                    punch_list = sorted(punch_list)

                    ws.cell(row, 8).value = ",".join(punch_list) + ","
                    ws.cell(row, 10).value = "Present"

                    total_updated += 1

            # Summary
            present = 0
            total = 0

            for r in range(2, ws.max_row + 1):

                reg = ws.cell(r, 2).value

                if reg:
                    total += 1

                    if ws.cell(r, 10).value == "Present":
                        present += 1

            absent = total - present

            percentage = round((present / total) * 100, 2) if total else 0

            # Footer
            last_row = ws.max_row + 2

            try:
                img = Image(LOGO_PATH)
                img.width = 120
                img.height = 80
                ws.add_image(img, f"A{last_row}")
            except:
                pass

            footer_row = last_row + 4

            ws.merge_cells(start_row=footer_row,
                           start_column=1,
                           end_row=footer_row,
                           end_column=6)

            title = ws.cell(footer_row, 1)
            title.value = "BIOSYNC Attendance Engine"
            title.font = Font(bold=True, size=12)
            title.alignment = Alignment(horizontal="center")

            ws.cell(footer_row + 1, 1).value = f"Class : {sh}"
            ws.cell(footer_row + 2, 1).value = f"Report Date : {FETCH_DATE_OBJ}"
            ws.cell(footer_row + 3, 1).value = f"Generated : {datetime.now()}"
            ws.cell(footer_row + 4, 1).value = f"Present : {present}"
            ws.cell(footer_row + 5, 1).value = f"Absent : {absent}"
            ws.cell(footer_row + 6, 1).value = f"Attendance % : {percentage}%"

        wb.save(ACTIVE_EXCEL_FILE)
        wb.close()

        compress_excel(ACTIVE_EXCEL_FILE)

        return total_updated

# ==================================================
# DEVICE WORKER
# ==================================================

def pull_device(device):

    name = device["name"]
    ip = device["ip"]
    port = device["port"]
    sheet = device["sheet"]

    if TARGET_DEVICES and name not in TARGET_DEVICES:
        return

    print(f"\n🔌 Connecting to {name}")

    if not is_ip_reachable(ip):

        print(f"❌ {name} OFFLINE")

        with tracker_lock:
            offline_devices.append(name)

        return

    zk = ZK(ip, port=port, timeout=10)

    try:

        conn = zk.connect()
        conn.disable_device()

        logs = conn.get_attendance()

        print(f"📥 {name} → {len(logs)} logs")

        filtered = [
            (rec.user_id, rec.timestamp.strftime("%H:%M"))
            for rec in logs
            if rec.timestamp.date() == FETCH_DATE_OBJ
        ]

        if not filtered:

            with tracker_lock:
                no_log_devices.append(name)

        else:

            updated = update_excel(sheet, filtered)

            with tracker_lock:
                successful_devices.append(name)

            print(f"✅ {name} updated {updated}")

        if AUTO_CLEAR_LOGS:
            conn.clear_attendance()

        conn.enable_device()
        conn.disconnect()

    except Exception as e:

        print(f"❌ {name} ERROR {e}")

        with tracker_lock:
            error_devices.append(name)

# ==================================================
# MAIN
# ==================================================

if __name__ == "__main__":

    print("\n🚀 BIOSYNC SDK ATTENDANCE PULL")
    print(f"📅 Fetch Date : {FETCH_DATE_OBJ}")
    print("=" * 60)

    threads = []

    for dev in DEVICES:
        t = threading.Thread(target=pull_device, args=(dev,))
        t.start()
        threads.append(t)

    for t in threads:
        t.join()

    print("\n" + "=" * 60)
    print("📊 EXECUTION SUMMARY")
    print("=" * 60)

    print(f"✅ Successful Devices : {len(successful_devices)}")
    print(f"❌ Offline Devices : {len(offline_devices)}")
    print(f"ℹ No Log Devices : {len(no_log_devices)}")
    print(f"❌ Error Devices : {len(error_devices)}")

    # Combine offline + error devices
    rerun_devices = offline_devices + error_devices

    if rerun_devices:

        print("\n⚠ DEVICES TO RERUN")

        for d in rerun_devices:
            print("→", d)

        formatted = "[" + ",".join(f'"{d}"' for d in rerun_devices) + "]"

        print("\n📋 COPY THIS FOR TARGET_DEVICES")
        print(formatted)

    print("\n🎉 SDK ATTENDANCE SYNC COMPLETED")
    print("=" * 60)